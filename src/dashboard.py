from openpyxl import load_workbook
from datetime import datetime


def update_dashboard(results, backtest=None):

    workbook = load_workbook("../dashboard/SwingTrade_Dashboard.xlsx")

    dashboard = workbook["Dashboard"]
    scanner = workbook["Scanner_Data"]

    # -------------------------
    # Dashboard Summary
    # -------------------------

    strong_buy = len([r for r in results if r["Signal"] == "Strong Buy"])
    buy = len([r for r in results if r["Signal"] == "Buy"])
    watch = len([r for r in results if r["Signal"] == "Watch"])

    dashboard["B3"] = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    dashboard["B5"] = len(results)
    dashboard["B6"] = strong_buy
    dashboard["B7"] = buy
    dashboard["B8"] = watch

    # -------------------------
    # Scanner Data
    # -------------------------

    scanner.delete_rows(2, scanner.max_row)

    row = 2

    for stock in results:

        scanner.cell(row=row, column=1).value = datetime.now().strftime("%d-%b-%Y")
        scanner.cell(row=row, column=2).value = stock["Symbol"]
        scanner.cell(row=row, column=3).value = stock["CMP"]
        scanner.cell(row=row, column=4).value = stock["Entry"]
        scanner.cell(row=row, column=5).value = stock["Target"]
        scanner.cell(row=row, column=6).value = stock["StopLoss"]
        scanner.cell(row=row, column=7).value = stock["Score"]
        scanner.cell(row=row, column=8).value = stock["Signal"]
        scanner.cell(row=row, column=9).value = stock["Reasons"]
        scanner.cell(row=row, column=10).value = stock["TradingView"]
        scanner.cell(row=row, column=11).value = stock["Screener"]

        row += 1

    # -------------------------
    # Backtest
    # -------------------------

    if backtest is not None:

        bt = workbook["Backtest"]

        bt["B2"] = backtest["Total Trades"]
        bt["B3"] = backtest["Winners"]
        bt["B4"] = backtest["Losers"]
        bt["B5"] = f'{backtest["Win Rate"]}%'
        bt["B6"] = f'{backtest["Avg Gain"]}%'
        bt["B7"] = f'{backtest["Avg Loss"]}%'
        bt["B8"] = backtest["Profit Factor"]
        bt["B9"] = backtest["Max Drawdown"]

    workbook.save("../dashboard/SwingTrade_Dashboard.xlsx")

    print("✅ Dashboard Updated")